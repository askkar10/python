# openpyxl
# pip install openpyxl

from openpyxl import load_workbook
excel_file = load_workbook('temp_data_01.xlsx')
results = []
ws = excel_file.worksheets[0]
for row in ws.iter_rows():
    results.append([cell.value for cell in row])

print(results)