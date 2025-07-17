# zeby obslugiwac forma excel trzeba zainstalowac zewnetrzny modul 'openpyxl'

# pip install openpyxl

from openpyxl import load_workbook
wb = load_workbook('temp_data_01.xlsx')
results = []
ws = wb.worksheets[0]
for row in ws.iter_rows():
  results.append([cell.value for cell in row])

print(results)